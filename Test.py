from openpyxl import load_workbook, Workbook
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tkinter as tk
from tkinter import filedialog
import os

# 创建临时弹窗用于选择文件
root = tk.Tk()
root.withdraw()
changed = False

# 让用户选择一个Excel文件
file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx *.xls")])
print("File selected 已选择文件: " + file_path)

def add_title_row(sheet, title, color3, font_size=24):
    # Unmerge all cells and store the merged ranges
    merged_ranges = list(sheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        sheet.unmerge_cells(str(merged_range))
    
    # Insert a new row at the top
    sheet.insert_rows(1)
    # Add the title to the first cell
    title_cell = sheet.cell(row=1, column=1, value=title)
    # Apply styles to the title cell
    title_cell.font = Font(bold=True, size=font_size)  # Set font size and bold
    title_cell.fill = PatternFill(start_color=color3, end_color=color3, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")  # Center the title text
    
    # Reapply the stored merged cell ranges, shifted down by one row
    for merged_range in merged_ranges:
        if ':' in str(merged_range):
            start_cell, end_cell = str(merged_range).split(':')
            start_col, start_row = openpyxl.utils.cell.coordinate_from_string(start_cell)
            end_col, end_row = openpyxl.utils.cell.coordinate_from_string(end_cell)
            new_start_cell = openpyxl.utils.cell.get_column_letter(openpyxl.utils.cell.column_index_from_string(start_col)) + str(start_row + 1)
            new_end_cell = openpyxl.utils.cell.get_column_letter(openpyxl.utils.cell.column_index_from_string(end_col)) + str(end_row + 1)
            new_range = f"{new_start_cell}:{new_end_cell}"
            sheet.merge_cells(new_range)

    # Merge the title row based on the number of columns with data
    merge_title_row(sheet)
    
    return sheet

def merge_title_row(sheet):
    # Determine the number of columns with data in the first row below the title
    max_col = sheet.max_column
    for col in range(1, max_col + 1):
        if sheet.cell(row=2, column=col).value is None:
            max_col = col - 1
            break
    
    # Merge the title row from the first column to the last column with data
    if max_col > 1:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        # Center the title text within the merged cell
        title_cell = sheet.cell(row=1, column=1)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True, size=title_cell.font.size)  # Ensure the font is bold and size is set

def save_sheets(sheet):
    if os.path.exists('output.xlsx'):
        output_wb = load_workbook('output.xlsx')
        output_sheet = output_wb.create_sheet('Processed Data')
        for row in sheet.rows:
            for cell in row:
                new_cell = output_sheet[cell.coordinate]
                new_cell.value = cell.value
                if cell.fill and cell.fill.fill_type:
                    new_cell.fill = PatternFill(
                        start_color=cell.fill.start_color.index,
                        end_color=cell.fill.end_color.index,
                        fill_type=cell.fill.fill_type
                    )
                if cell.font:
                    new_cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        vertAlign=cell.font.vertAlign,
                        underline=cell.font.underline,
                        strike=cell.font.strike,
                        color=cell.font.color
                    )
                if cell.alignment:
                    new_cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        text_rotation=cell.alignment.text_rotation,
                        wrap_text=cell.alignment.wrap_text,
                        shrink_to_fit=cell.alignment.shrink_to_fit,
                        indent=cell.alignment.indent
                    )
        for merged_range in sheet.merged_cells.ranges:
            output_sheet.merge_cells(str(merged_range))
        output_wb.save('test.xlsx')
        print("Output file saved successfully")
    else:
        sheet.save('test.xlsx')

if __name__ == '__main__':
    wb = load_workbook(file_path)
    font_size = 24
    sheet = wb.active
    title = "Processed Data"
    color3 = "FFC000"  # Hex color code for light orange
    sheet = add_title_row(sheet, title, color3, font_size)
    save_sheets(sheet)