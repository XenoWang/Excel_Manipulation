from openpyxl import load_workbook,Workbook
import openpyxl
from openpyxl.styles import Font, PatternFill, Color
import tkinter as tk
from tkinter import filedialog
import os

# 创建临时弹窗用于选择文件
root = tk.Tk()
root.withdraw()
changed = False

# 让用户选择一个Excel文件
file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx *.xls")])
print("File selected 已选择文件: " + file_path)

def merge_cells(sheet):
    global changed
    # Unmerge all cells to avoid bugs 解除现有合并单元格
    for merged_range in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(merged_range))

    # Ask user for the column number to stop merging 询问用户要停止合并的列
    column_num = int(input("Enter the column number to stop merging (e.g. 1 for A, 2 for B, etc.): "))

    # Merge cells in the first column up to the user-inputted column number 合并单元格至用户输入的列
    current_value = None
    start_row = None
    for row_idx, row in enumerate(sheet.rows, start=1):
        cell_value = row[0].value
        if cell_value != current_value:
            if start_row is not None:
                # Merge cells for the previous group 合并单元格于上一个组
                for col_idx in range(1, column_num + 1):
                    sheet.merge_cells(start_row=start_row, start_column=col_idx, end_row=row_idx - 1, end_column=col_idx)
            current_value = cell_value
            start_row = row_idx

    # Merge cells for the last group 合并单元格于最后一个组
    for col_idx in range(1, column_num + 1):
        sheet.merge_cells(start_row=start_row, start_column=col_idx, end_row=sheet.max_row, end_column=col_idx)
    changed = True
    return sheet

def add_title_row(sheet, title,color3):
    # Unmerge all cells and store the merged ranges
    merged_ranges = list(sheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        sheet.unmerge_cells(str(merged_range))
    
    # Insert a new row at the top
    sheet.insert_rows(1)
    # Add the title to the first cell
    sheet.cell(row=1, column=1, value=title)
    # Apply styles to the title cell
    title_cell = sheet.cell(row=1, column=1)
    title_cell.font = Font(bold=True)
    title_cell.fill = PatternFill(start_color=color3, end_color=color3, fill_type="solid")
    
    # Reapply the stored merged cell ranges, shifted down by one row
    for merged_range in merged_ranges:
        new_range = str(merged_range).replace('1:', '2:').replace('1,', '2,')
        sheet.merge_cells(new_range)
    return sheet

def save_sheets(sheet):
    if os.path.exists('output.xlsx'):
        output_wb = load_workbook('output.xlsx')
        output_sheet = output_wb.create_sheet('Processed Data')
        for row in sheet.rows:
            for cell in row:
                # Copy the value to the new sheet and save it 复制值并保存
                new_cell = output_sheet[cell.coordinate]
                new_cell.value = cell.value
                # Create a new PatternFill object for the new cell 创建新的PatternFill对象
                if cell.fill and cell.fill.fill_type:
                    new_cell.fill = PatternFill(
                        start_color=cell.fill.start_color.index,
                        end_color=cell.fill.end_color.index,
                        fill_type=cell.fill.fill_type
                    )
        # Manually merge cells 手动合并
        for merged_range in sheet.merged_cells.ranges:
            output_sheet.merge_cells(str(merged_range))
        output_wb.save('output.xlsx')
        print("Output file saved successfully")
    else:
        output_wb = load_workbook(file_path)
        output_sheet = output_wb.active
        for row in sheet.rows:
            for cell in row:
                # Copy the value to the new sheet 复制值
                new_cell = output_sheet[cell.coordinate]
                new_cell.value = cell.value
                # Create a new PatternFill object for the new cell 创建新的PatternFill对象
                if cell.fill and cell.fill.fill_type:
                    new_cell.fill = PatternFill(
                        start_color=cell.fill.start_color.index,
                        end_color=cell.fill.end_color.index,
                        fill_type=cell.fill.fill_type
                    )
        # Manually merge cells 手动合并
        for merged_range in sheet.merged_cells.ranges:
            output_sheet.merge_cells(str(merged_range))
        output_wb.save('output.xlsx')
        print("Output file saved successfully")

def remove_color_format(sheet, row_number):
    for cell in sheet[row_number]:
        cell.fill = None

def change_title_color(sheet, row_number,color):
    global changed
    print("Changing row color...")
    # Create a Fill object with the desired color 创建填充对象
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    # Change background color of specified row 更改指定行的背景颜色
    for cell in sheet[row_number]:
        cell.fill = fill
    changed = True
    return sheet

def change_row_colors(sheet, color1, color2):
    global changed
    ws = sheet
    # Get the maximum row number of the worksheet 返回工作表的最大行号
    max_row = ws.max_row

    # Mark the rows with color 给行打上颜色
    row_count = 0
    current_row = 2  # Start from the second row 从第二行开始
    while current_row <= max_row:
        # Check if there are merged cells in the current row 查询当前行是否有合并单元格
        merged_cell_ranges = ws.merged_cells.ranges
        intersecting_ranges = [rng for rng in merged_cell_ranges if rng.min_row <= current_row and rng.max_row >= current_row]
        if intersecting_ranges:
            # If there are merged cells, consider them as one row 如果有合并单元格，将它们视为一个行
            row_height = max(rng.max_row for rng in intersecting_ranges) - min(rng.min_row for rng in intersecting_ranges) + 1
            for row in range(current_row, current_row + row_height):
                for cell in ws.iter_rows(min_row=row, max_row=row, values_only=False):
                    for c in cell:
                        if row_count % 2 == 0:
                            c.fill = openpyxl.styles.PatternFill(start_color=color1, end_color=color1, fill_type='solid')
                        else:
                            c.fill = openpyxl.styles.PatternFill(start_color=color2, end_color=color2, fill_type='solid')
            row_count += 1
            current_row += row_height
        else:
            # If there are no merged cells, consider each row separately 如果没有合并单元格，将每个行单独考虑
            for cell in ws.iter_rows(min_row=current_row, max_row=current_row, values_only=False):
                for c in cell:
                    if row_count % 2 == 0:
                        c.fill = openpyxl.styles.PatternFill(start_color=color1, end_color=color1, fill_type='solid')
                    else:
                        c.fill = openpyxl.styles.PatternFill(start_color=color2, end_color=color2, fill_type='solid')
            row_count += 1
            current_row += 1
    changed = True
    return wb

# Load the selected Excel file 加载选择的Excel文件
wb = load_workbook(file_path, read_only=False)
sheet = wb.active
while True:
    print("Menu 菜单:")
    print("1. Merge cells 合并单元格")
    print("2. Change background color of cells 更改单元格的背景颜色")
    print("0. Quit the program 退出程序")
    choice = input("Enter your choice 请输入你的选择: ")

    if choice == "1":
        if file_path:
            # 加载Excel文件
            sheet = merge_cells(sheet)
    elif choice == "2":
        if file_path:
            while True:
                print("Menu:")
                print("1. Change title color 更改标题颜色")
                print("2. Change row colors 更改行颜色")
                print("0. Go back to previous menu 返回上一级菜单")
                choice = input("Enter your choice 请输入你的选择: ")

                if choice == "1":
                    row_number = 1
                    color = input("Enter the color hex code 请输入颜色的十六进制代码 (e.g. FF0000 for red no need to add #, 不需要加#): ")
                    sheet = change_title_color(sheet, row_number,color)
                    break
                elif choice == "2":
                    # Ask user for two colors 询问用户两个颜色
                    color1 = input("Enter the color1 hex code 请输入颜色1的十六进制代码 (e.g. FF0000 for red no need to add #, 不需要加#): ")
                    color2 = input("Enter the color2 hex code 请输入颜色2的十六进制代码 (e.g. FF0000 for red no need to add #, 不需要加#): ")
                    change_row_colors(sheet, color1, color2)
                    break
                elif choice == "0":
                    break
                else:
                    print("Invalid choice. Please try again. 无效的选择，请重试。")
        else:
            print("No file selected 未选择文件")
    elif choice == "0":
        if changed:
            save_sheets(sheet)
        else:
            print("No changes made, not saving to output.xlsx file. 未进行任何更改,不保存到output.xlsx文件中。")
        break
    else:
        print("Invalid choice. Please try again. 无效的选择，请重试。")